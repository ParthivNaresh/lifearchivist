"""Utility functions for extract tools."""

from datetime import datetime
from typing import Any, Callable, Dict, List, Optional

from openpyxl.cell import Cell


class PDFMetadataExtractor:
    """Utility class for extracting PDF metadata."""

    @staticmethod
    def extract_date_field(
        pdf_metadata,
        field_key: str,
        parse_func: Callable[[Optional[str]], Optional[str]],
    ) -> Optional[str]:
        """
        Extract and parse date field from PDF metadata.

        Args:
            pdf_metadata: PDF metadata object
            field_key: Key to extract (e.g., "/CreationDate")
            parse_func: Function to parse the date

        Returns:
            Parsed date string or None
        """
        if field_key not in pdf_metadata:
            return None

        raw_value = pdf_metadata.get(field_key)
        if raw_value is None:
            return None

        result = parse_func(str(raw_value))
        return result if result is not None else None

    @staticmethod
    def extract_text_fields(
        pdf_metadata, field_mapping: Dict[str, str]
    ) -> Dict[str, str]:
        """
        Extract text fields from PDF metadata.

        Args:
            pdf_metadata: PDF metadata object
            field_mapping: Mapping of PDF keys to metadata keys

        Returns:
            Dictionary of extracted text fields
        """
        result = {}

        for pdf_key, metadata_key in field_mapping.items():
            if pdf_key in pdf_metadata:
                value = pdf_metadata[pdf_key]
                if value and isinstance(value, str):
                    result[metadata_key] = value.strip()

        return result

    @staticmethod
    def create_metadata_dict(
        created_date: Optional[str],
        modified_date: Optional[str],
        text_fields: Dict[str, str],
    ) -> Dict[str, Any]:
        """
        Create metadata dictionary from extracted fields.

        Args:
            created_date: Creation date string
            modified_date: Modification date string
            text_fields: Dictionary of text fields

        Returns:
            Combined metadata dictionary
        """
        metadata: Dict[str, Any] = {}

        if created_date:
            metadata["document_created_at"] = created_date

        if modified_date:
            metadata["document_modified_at"] = modified_date

        metadata.update(text_fields)

        return metadata


class CSVCellFormatter:
    """Utility class for formatting CSV cell values."""

    @staticmethod
    def try_format_as_date(cell_value: str) -> str:
        """
        Try to format cell value as date.

        Args:
            cell_value: Cell value string

        Returns:
            Formatted date string or original value
        """
        from dateutil import parser as date_parser  # type: ignore[import-untyped]

        if (
            not cell_value
            or cell_value.replace(".", "").replace(",", "").replace("-", "").isdigit()
        ):
            return cell_value

        try:
            parsed_date = date_parser.parse(cell_value, fuzzy=False)
            if not cell_value.isdigit():
                return str(parsed_date.strftime("%Y-%m-%d"))
        except (ValueError, TypeError):
            pass

        return cell_value

    @staticmethod
    def try_format_as_number(cell_value: str) -> str:
        """
        Try to format cell value as number with thousand separators.

        Args:
            cell_value: Cell value string

        Returns:
            Formatted number string or original value
        """
        if not cell_value.replace(",", "").replace(".", "").replace("-", "").isdigit():
            return cell_value

        try:
            clean_num = cell_value.replace(",", "")
            if "." in clean_num:
                num = float(clean_num)
                if num.is_integer():
                    return f"{int(num):,}"
                return f"{num:,.2f}"
            else:
                num = int(clean_num)
                if abs(num) >= 1000:
                    return f"{num:,}"
        except ValueError:
            pass

        return cell_value

    @staticmethod
    def format_cell(cell: str) -> str:
        """
        Format a CSV cell value.

        Args:
            cell: Raw cell value

        Returns:
            Formatted cell value
        """
        cell_value = cell.strip()
        cell_value = CSVCellFormatter.try_format_as_date(cell_value)
        cell_value = CSVCellFormatter.try_format_as_number(cell_value)
        return cell_value


class CSVProcessor:
    """Utility class for processing CSV files."""

    @staticmethod
    async def read_csv_content(file_path, encoding: str) -> str:
        """
        Read CSV file content with encoding.

        Args:
            file_path: Path to CSV file
            encoding: Encoding to use

        Returns:
            File content string
        """
        import aiofiles

        try:
            async with aiofiles.open(
                file_path, "r", encoding=encoding, errors="replace"
            ) as f:
                return await f.read()
        except UnicodeDecodeError:
            async with aiofiles.open(
                file_path, "r", encoding="latin-1", errors="replace"
            ) as f:
                return await f.read()

    @staticmethod
    def process_csv_row(row: List[str]) -> List[str]:
        """
        Process and clean CSV row.

        Args:
            row: Raw row values

        Returns:
            Cleaned row values
        """
        return [CSVCellFormatter.format_cell(cell) for cell in row]

    @staticmethod
    def format_header_row(cleaned_row: List[str]) -> List[str]:
        """
        Format header row.

        Args:
            cleaned_row: Cleaned row values

        Returns:
            Formatted header lines
        """
        header_text = " | ".join(cleaned_row)
        return [f"[CSV HEADERS] {header_text}", "-" * 80]

    @staticmethod
    def format_data_row(cleaned_row: List[str]) -> str:
        """
        Format data row.

        Args:
            cleaned_row: Cleaned row values

        Returns:
            Formatted row string
        """
        return " | ".join(cleaned_row)

    @staticmethod
    def create_summary(
        row_count: int,
        delimiter: str,
        encoding: str,
        header_row: Optional[List[str]],
    ) -> List[str]:
        """
        Create CSV summary.

        Args:
            row_count: Number of rows processed
            delimiter: Delimiter used
            encoding: Encoding used
            header_row: Header row if present

        Returns:
            Summary lines
        """
        summary = [
            "",
            "[CSV SUMMARY]",
            f"Total rows processed: {row_count}",
            f"Delimiter used: '{delimiter}'",
            f"Encoding: {encoding}",
        ]

        if header_row:
            summary.append(f"Number of columns: {len(header_row)}")

        return summary


class OCRProcessor:
    """Utility class for OCR processing."""

    @staticmethod
    def get_num_pages(image) -> int:
        """
        Get number of pages in image.

        Args:
            image: PIL Image

        Returns:
            Number of pages
        """
        if hasattr(image, "n_frames"):
            return int(image.n_frames)
        return 1

    @staticmethod
    def calculate_avg_confidence(confidences: List[int]) -> float:
        """
        Calculate average confidence from list.

        Args:
            confidences: List of confidence scores

        Returns:
            Average confidence
        """
        if not confidences:
            return 0.0
        return sum(confidences) / len(confidences)

    @staticmethod
    def get_confidence_level(avg_confidence: float) -> str:
        """
        Get confidence level label.

        Args:
            avg_confidence: Average confidence score

        Returns:
            Confidence level string
        """
        if avg_confidence > 80:
            return "High"
        elif avg_confidence > 60:
            return "Medium"
        return "Low"

    @staticmethod
    async def process_single_page(
        image,
        page_num: int,
        num_pages: int,
        preprocess_func,
        ocr_config: str = r"--oem 3 --psm 3",
    ) -> tuple[Optional[str], List[int]]:
        """
        Process a single page with OCR.

        Args:
            image: PIL Image
            page_num: Page number (0-indexed)
            num_pages: Total number of pages
            preprocess_func: Function to preprocess image
            ocr_config: Tesseract configuration

        Returns:
            Tuple of (extracted_text, confidence_scores)
        """
        import asyncio
        import logging

        import pytesseract

        if num_pages > 1:
            image.seek(page_num)
            logging.info(f"Processing page {page_num + 1} of {num_pages}")

        processed_image = preprocess_func(image.copy())

        try:
            text = await asyncio.to_thread(
                pytesseract.image_to_string, processed_image, config=ocr_config
            )

            data = await asyncio.to_thread(
                pytesseract.image_to_data,
                processed_image,
                output_type=pytesseract.Output.DICT,
                config=ocr_config,
            )

            confidences = [conf for conf in data["conf"] if conf > 0]
            avg_confidence = OCRProcessor.calculate_avg_confidence(confidences)

            if text.strip():
                confidence_level = OCRProcessor.get_confidence_level(avg_confidence)
                logging.info(
                    f"OCR confidence for page {page_num + 1}: {confidence_level} ({avg_confidence:.1f}%)"
                )
                return text.strip(), confidences

        except pytesseract.TesseractNotFoundError as e:
            raise ValueError(
                "Tesseract OCR is not installed. Please install it using: "
                "brew install tesseract (macOS) or apt-get install tesseract-ocr (Linux)"
            ) from e
        except Exception as ocr_error:
            logging.warning(f"OCR failed for page {page_num + 1}: {ocr_error}")

        return None, []

    @staticmethod
    def format_multipage_text(all_text: List[str], num_pages: int) -> List[str]:
        """
        Format text for multi-page documents.

        Args:
            all_text: List of text from pages
            num_pages: Total number of pages

        Returns:
            Formatted text list with separators
        """
        if num_pages <= 1 or not all_text:
            return all_text

        formatted = [all_text[0]]
        for i, text in enumerate(all_text[1:], start=2):
            formatted.append(f"\n--- Page {i} ---\n")
            formatted.append(text)

        return formatted

    @staticmethod
    def log_ocr_summary(
        num_pages: int, all_text: List[str], total_confidence: List[int]
    ) -> None:
        """
        Log OCR summary statistics.

        Args:
            num_pages: Number of pages processed
            all_text: List of extracted text
            total_confidence: List of all confidence scores
        """
        import logging

        if all_text:
            word_count = sum(len(text.split()) for text in all_text)
            overall_confidence = OCRProcessor.calculate_avg_confidence(total_confidence)
            logging.info(
                f"OCR complete: {num_pages} pages, {word_count} words, {overall_confidence:.1f}% avg confidence"
            )
        else:
            logging.warning("No text detected in image")


class ExcelSheetProcessor:
    """Utility class for processing Excel sheets."""

    @staticmethod
    def is_empty_sheet(sheet) -> bool:
        """
        Check if sheet is empty.

        Args:
            sheet: Excel worksheet

        Returns:
            True if empty
        """
        return bool(sheet.max_row == 0 or sheet.max_column == 0)

    @staticmethod
    def create_sheet_header(
        sheet_name: str, max_row: int, max_column: int
    ) -> List[str]:
        """
        Create sheet header content.

        Args:
            sheet_name: Name of the sheet
            max_row: Maximum row count
            max_column: Maximum column count

        Returns:
            List of header lines
        """
        return [
            f"\n[SHEET: {sheet_name}]",
            f"[DIMENSIONS: {max_row} rows × {max_column} columns]",
        ]

    @staticmethod
    def format_headers(headers: List[str]) -> List[str]:
        """
        Format header row.

        Args:
            headers: List of header values

        Returns:
            List of formatted header lines
        """
        header_text = " | ".join(
            h if h else f"Column {i+1}" for i, h in enumerate(headers)
        )
        return [f"[HEADERS] {header_text}", "-" * 80]

    @staticmethod
    def extract_row_values(sheet, row_idx: int, format_func) -> tuple[List[str], bool]:
        """
        Extract values from a row.

        Args:
            sheet: Excel worksheet
            row_idx: Row index
            format_func: Function to format cell values

        Returns:
            Tuple of (row_values, has_content)
        """
        row_values = []
        has_content = False

        for col_idx in range(1, sheet.max_column + 1):
            cell = sheet.cell(row=row_idx, column=col_idx)
            value = format_func(cell)
            row_values.append(value)
            if value:
                has_content = True

        return row_values, has_content

    @staticmethod
    def process_data_rows(
        sheet,
        start_row: int,
        format_func,
        max_rows: int = 10000,
        max_empty_rows: int = 5,
    ) -> List[str]:
        """
        Process data rows from sheet.

        Args:
            sheet: Excel worksheet
            start_row: Starting row index
            format_func: Function to format cell values
            max_rows: Maximum rows to extract
            max_empty_rows: Maximum consecutive empty rows before stopping

        Returns:
            List of formatted data rows
        """
        data_rows = []
        empty_row_count = 0

        for row_idx in range(start_row, sheet.max_row + 1):
            row_values, has_content = ExcelSheetProcessor.extract_row_values(
                sheet, row_idx, format_func
            )

            if has_content:
                empty_row_count = 0
                row_text = " | ".join(row_values)
                data_rows.append(row_text)

                if len(data_rows) >= max_rows:
                    data_rows.append(f"[... truncated after {max_rows} rows ...]")
                    break
            else:
                empty_row_count += 1
                if empty_row_count >= max_empty_rows:
                    break

        return data_rows

    @staticmethod
    def process_single_sheet(
        sheet, sheet_name: str, format_func, header_detector
    ) -> List[str]:
        """
        Process a single Excel sheet.

        Args:
            sheet: Excel worksheet
            sheet_name: Name of the sheet
            format_func: Function to format cell values
            header_detector: Function to detect headers

        Returns:
            List of content lines for the sheet
        """
        if ExcelSheetProcessor.is_empty_sheet(sheet):
            return []

        sheet_content = ExcelSheetProcessor.create_sheet_header(
            sheet_name, sheet.max_row, sheet.max_column
        )

        header_row_idx, headers = header_detector(sheet)

        if headers and any(headers):
            sheet_content.extend(ExcelSheetProcessor.format_headers(headers))

        start_row = header_row_idx + 1 if header_row_idx > 0 else 1
        data_rows = ExcelSheetProcessor.process_data_rows(sheet, start_row, format_func)

        if data_rows:
            sheet_content.extend(data_rows)
        else:
            sheet_content.append("[No data rows found]")

        return sheet_content


class CellValueFormatter:
    """Utility class for formatting Excel cell values."""

    @staticmethod
    def format_datetime(value: datetime) -> str:
        """
        Format datetime value.

        Args:
            value: Datetime object

        Returns:
            Formatted datetime string
        """
        if value.time() == datetime.min.time():
            return value.strftime("%Y-%m-%d")
        return value.strftime("%Y-%m-%d %H:%M:%S")

    @staticmethod
    def format_boolean(value: bool) -> str:
        """
        Format boolean value.

        Args:
            value: Boolean value

        Returns:
            Formatted boolean string
        """
        return "TRUE" if value else "FALSE"

    @staticmethod
    def extract_currency_symbol(format_str: str) -> Optional[str]:
        """
        Extract currency symbol from format string.

        Args:
            format_str: Excel number format string

        Returns:
            Currency symbol or None
        """
        currency_symbols = ["$", "€", "£", "¥"]
        for symbol in currency_symbols:
            if symbol in format_str:
                return symbol
        return None

    @staticmethod
    def format_currency(value: float, symbol: str) -> str:
        """
        Format currency value.

        Args:
            value: Numeric value
            symbol: Currency symbol

        Returns:
            Formatted currency string
        """
        return f"{symbol}{value:,.2f}"

    @staticmethod
    def format_percentage(value: float) -> str:
        """
        Format percentage value.

        Args:
            value: Numeric value (0-1 range)

        Returns:
            Formatted percentage string
        """
        return f"{value * 100:.2f}%"

    @staticmethod
    def format_number(value: float) -> str:
        """
        Format numeric value.

        Args:
            value: Numeric value

        Returns:
            Formatted number string
        """
        if isinstance(value, float) and value.is_integer():
            return str(int(value))
        return str(value)

    @staticmethod
    def format_large_integer(value: int) -> str:
        """
        Format large integer with thousand separators.

        Args:
            value: Integer value

        Returns:
            Formatted integer string
        """
        return f"{value:,}"

    @staticmethod
    def is_percentage_format(cell: Cell) -> bool:
        """
        Check if cell has percentage format.

        Args:
            cell: Excel cell

        Returns:
            True if percentage format
        """
        return cell.number_format and "%" in str(cell.number_format)

    @staticmethod
    def is_currency_format(cell: Cell) -> bool:
        """
        Check if cell has currency format.

        Args:
            cell: Excel cell

        Returns:
            True if currency format
        """
        if not cell.number_format:
            return False
        format_str = str(cell.number_format)
        return any(symbol in format_str for symbol in ["$", "€", "£", "¥"])

    @staticmethod
    def format_numeric_value(cell: Cell, value: float) -> str:
        """
        Format numeric value based on cell format.

        Args:
            cell: Excel cell
            value: Numeric value

        Returns:
            Formatted string
        """
        if CellValueFormatter.is_percentage_format(cell):
            return CellValueFormatter.format_percentage(value)

        if CellValueFormatter.is_currency_format(cell):
            format_str = str(cell.number_format)
            symbol = CellValueFormatter.extract_currency_symbol(format_str)
            if symbol:
                return CellValueFormatter.format_currency(value, symbol)
            return f"{value:,.2f}"

        if isinstance(value, int) and abs(value) >= 1000:
            return CellValueFormatter.format_large_integer(value)

        return CellValueFormatter.format_number(value)

    @staticmethod
    def format_formula_result(cell: Cell, value: Any) -> str:
        """
        Format formula result.

        Args:
            cell: Excel cell
            value: Formula result value

        Returns:
            Formatted string
        """
        if value is not None:
            return str(value)
        return "[Formula Error]"
